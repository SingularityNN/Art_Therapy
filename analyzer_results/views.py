from django.http import HttpRequest, HttpResponse
from django.shortcuts import render, redirect, reverse, get_object_or_404
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from image_analyzer.analyzer import analyze_image
from .forms import ExperimentForm, NoteForm
from .models import Experiments
import os
import re
from django.conf import settings
from django.core.files import File

def import_drawings(request: HttpRequest) -> HttpResponse:
    """Импортирует все пары картинок из папки art, выполняет анализ и сохраняет в БД"""
    if request.method != 'POST':
        return redirect('analyzer_results:experiments_list')

    # Путь к папке art (рядом с manage.py)
    art_folder = os.path.join(settings.BASE_DIR, 'art')
    if not os.path.isdir(art_folder):
        messages.error(request, f'Папка art не найдена: {art_folder}')
        return redirect('analyzer_results:experiments_list')

    # Регулярное выражение для парсинга имён файлов
    # Ожидаем формат: art_<идентификатор>_1.jpeg (или _2, любое расширение)
    pattern = re.compile(r'art_(.+)_([12])\.[a-zA-Z0-9]+$')

    # Словарь: идентификатор -> {1: (имя_файла, полный_путь), 2: ...}
    pairs = {}

    for filename in os.listdir(art_folder):
        full_path = os.path.join(art_folder, filename)
        if not os.path.isfile(full_path):
            continue

        match = pattern.match(filename)
        if not match:
            continue  # пропускаем файлы, не соответствующие шаблону

        base_id = match.group(1)  # например, "aksenova" или "chugrova_32f"
        number = int(match.group(2))

        if base_id not in pairs:
            pairs[base_id] = {}
        pairs[base_id][number] = (filename, full_path)

    # Статистика
    created = 0
    updated = 0
    skipped = 0

    for exp_id, files in pairs.items():
        # Проверяем наличие обоих номеров
        if 1 not in files or 2 not in files:
            skipped += 1
            continue

        filename1, path1 = files[1]
        filename2, path2 = files[2]

        # Относительные пути для функции analyze_image (как в ручной функции)
        rel_path1 = os.path.join('art', filename1)
        rel_path2 = os.path.join('art', filename2)

        try:
            # Выполняем анализ
            analyzer_res_first = analyze_image(rel_path1)
            analyzer_res_second = analyze_image(rel_path2)
        except Exception as e:
            messages.error(request, f'Ошибка анализа для "{exp_id}": {e}')
            skipped += 1
            continue

        # Получаем существующую запись или создаём новую
        obj, created_flag = Experiments.objects.get_or_create(id=exp_id)

        # Если запись уже существовала, удаляем старые файлы
        if not created_flag:
            if obj.drawing_first and os.path.exists(obj.drawing_first.path):
                os.remove(obj.drawing_first.path)
            if obj.drawing_second and os.path.exists(obj.drawing_second.path):
                os.remove(obj.drawing_second.path)

        # Сохраняем новые изображения в ImageField
        with open(path1, 'rb') as f1, open(path2, 'rb') as f2:
            obj.drawing_first.save(filename1, File(f1), save=False)
            obj.drawing_second.save(filename2, File(f2), save=False)

        # Заполняем JSON-поля результатами анализа
        obj.analyzer_res_first_json = analyzer_res_first
        obj.analyzer_res_second_json = analyzer_res_second
        obj.save()

        if created_flag:
            created += 1
        else:
            updated += 1

    messages.success(
        request,
        f'Импорт завершён. Создано: {created}, обновлено: {updated}, пропущено (неполные пары): {skipped}'
    )
    return redirect('analyzer_results:experiments_list')


def export_to_xlsx(request: HttpRequest) -> HttpResponse:
    queryset = Experiments.objects.all()
    keys = ['ignored_white_pixels',
            'avg_hue',
            'avg_saturation',
            'avg_brightness',
            'red_percent',
            'green_percent',
            'blue_percent']

    wb = Workbook()
    ws = wb.active
    ws.title = "drawings_data"
    headers = ['ID',
            'ignored_white_pixels (%)',
            'avg_hue (0-179)',
            'avg_saturation (0-255)',
            'avg_brightness (0-255)',
            'red_percent (%)',
            'green_percent (%)',
            'blue_percent (%)']
    ws.append(headers)

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18

    header_fill = PatternFill(start_color="D3D3D3",
                              end_color="D3D3D3",
                              fill_type="solid")

    for cell in ws[1]:  # ws[1] — это первая строка
        cell.fill = header_fill


    for obj in queryset:
        row = [obj.id + '_1']
        for key in keys:
            value = obj.analyzer_res_first_json.get(key, '')
            row.append(value)
        ws.append(row)

        row = [obj.id + '_2']
        for key in keys:
            value = obj.analyzer_res_second_json.get(key, '')
            row.append(value)
        ws.append(row)

        columns_format = {
            'B': '0.00',
            'F': '0.00',
            'G': '0.00',
            'H': '0.00',
        }

        for row in range(2, ws.max_row + 1):
            for col_letter, fmt in columns_format.items():
                cell = ws[f"{col_letter}{row}"]
                # Если значение в ячейке - число (int или float), применяем формат
                if isinstance(cell.value, (int, float)):
                    cell.number_format = fmt


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="export_drawings_data.xlsx"'
    wb.save(response)
    return response

def add_experiment(request: HttpRequest) -> HttpResponse:
    if request.method == 'POST':
        form = ExperimentForm(request.POST, request.FILES)
        if form.is_valid():

            # get experiment.drawing_first
            drawing_first_file = form.cleaned_data['drawing_first']

            # get experiment.drawing_second
            drawing_second_file = form.cleaned_data['drawing_second']

            # get experiment.id
            file_name_first = drawing_first_file.name
            file_name_second = drawing_second_file.name
            file_name_shortened = file_name_first[4:-7]

            if Experiments.objects.filter(id=file_name_shortened).exists():
                messages.error(request, f'Эксперимент с id "{file_name_shortened}" уже существует')
                return render(request, 'analyzer_results/add_experiment.html', {'form': form})

            # get experiment.analyzer_res_first_json
            image_path_first = "art/" + file_name_first
            analyzer_res_first = analyze_image(image_path_first)

            # get experiment.analyzer_res_second_json
            image_path_second = "art/" + file_name_second
            analyzer_res_second = analyze_image(image_path_second)

            experiment = Experiments(
                id = file_name_shortened,
                drawing_first = drawing_first_file,
                analyzer_res_first_json = analyzer_res_first,
                drawing_second = drawing_second_file,
                analyzer_res_second_json = analyzer_res_second,
            )

            experiment.save()

            messages.success(request, f'Эксперимент "{experiment.id}" успешно добавлен.')
            return redirect(reverse('analyzer_results:experiments_list'))

        else:
            messages.error(request, 'Исправьте ошибки в форме.')

    else:
        form = ExperimentForm()

    return render(request, 'analyzer_results/add_experiment.html', {'form': form})

def experiment_detail(request: HttpRequest, id) -> HttpResponse:
    exp = get_object_or_404(Experiments, id=id)

    if request.method == 'POST':
        form = NoteForm(request.POST, instance=exp)
        if form.is_valid():
            form.save()
            return redirect(reverse('analyzer_results:experiments_list'))

        else:
            messages.error(request, 'Исправьте ошибки в форме.')

    else:
        form = NoteForm(instance=exp)
    return render(request, 'analyzer_results/experiment_detail.html', {'form': form, 'exp': exp,})

def delete_experiment (request: HttpRequest, id) -> HttpResponse:
    obj = Experiments.objects.get(id=id)
    obj.delete()
    return redirect(reverse('analyzer_results:experiments_list'))

def start_page(request: HttpRequest) -> HttpResponse:
    return render(request, 'analyzer_results/start_page.html')

def experiments_list_all(request: HttpRequest) -> HttpResponse:
    context = {
        "experiments": Experiments.objects.all(),
    }
    return render(request, 'analyzer_results/experiments_list.html', context=context)
